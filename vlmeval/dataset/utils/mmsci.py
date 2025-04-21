# flake8: noqa
import evaluate
from tqdm import tqdm
import numpy as np
import os.path as osp
import pandas as pd
from vlmeval import load, dump, track_progress_rich

g_eval_prompt = """
You will be given a oracle caption that describes an image. You will then be given a second caption written for the same image.

Your task is to rate the second caption on one metric.

Evaluation Criteria:

Relevance (1-5) - The extent to which the second caption is relevant to the key elements and context described in the oracle caption. A relevant caption should focus on the same subjects, objects, actions, or context highlighted in the oracle caption, without introducing unrelated or extraneous details.

Evaluation Steps:

1. Review the Oracle Caption: Carefully read the oracle caption to understand the main elements and context it describes.
2. Review the Second Caption: Assess whether the second caption focuses on the same key elements and context as the oracle caption. Evaluate if the second caption stays on topic and does not introduce irrelevant details.
3. Assign a Score for Relevance: Based on the Evaluation Criteria, rate how relevant the second caption is to the oracle caption's description of the same image.
"""

generate_atomic_facts_sys_prompt = """
You will be given a caption for a figure containing multiple panels, which includes descriptions for the entire figure as well as each individual panel. Your task is to break down the caption into separate, independent descriptions for the entire figure and each panel, formatted appropriately and separated by '-'
"""

generate_atomic_facts_usr_prompt_one_shot = """
The figure consists of four sub-figures labeled a, b, c, and d. All four images appear to be scanning electron microscope (SEM) images showing the microstructure of different materials, likely related to the iron-based cathode catalysts described in the article.\n\na. This image shows a highly porous structure with interconnected particles forming a network. The particles appear to be in the nanometer to micrometer size range. The scale bar indicates 1 \u03bcm.\n\nb. This image displays a closer view of what seems to be a similar material to (a), but at a higher magnification. The individual particles are more clearly visible, showing a rough, granular texture. The scale bar indicates 200 nm.\n\nc. This image reveals a different morphology compared to (a) and (b). It shows larger, more consolidated structures with a rougher surface texture. There are still visible pores and gaps between the structures. The scale bar indicates 1 \u03bcm.\n\nd. This final image appears to be a cross-sectional view of a porous material, possibly showing the internal structure of the catalyst. It reveals a highly interconnected network of pores and channels throughout the material. The scale bar indicates 200 nm.\n\nThese images likely represent different stages or variations of the iron-acetate/phenanthroline/zeolitic-imidazolate-framework-derived electrocatalyst mentioned in the article. The varying structures and porosities shown in these images could be related to the enhanced mass-transport properties and increased volumetric activity described in the text.
Model_Response:
- The figure consists of four sub-figures labeled a, b, c, and d.
- All four images appear to be scanning electron microscope (SEM) images.
- The images show the microstructure of different materials.
- The materials are likely related to the iron-based cathode catalysts described in the article.
- Image a shows a highly porous structure with interconnected particles forming a network.
- The particles in image a are in the nanometer to micrometer size range.
- The scale bar in image a indicates 1 μm.
- Image b displays a closer view of a material similar to that in image a but at higher magnification.
- The individual particles in image b are more clearly visible and show a rough, granular texture.
- The scale bar in image b indicates 200 nm.
- Image c shows larger, more consolidated structures with a rougher surface texture.
- There are visible pores and gaps between the structures in image c.
- The scale bar in image c indicates 1 μm.
- Image d appears to be a cross-sectional view of a porous material.
- Image d reveals the internal structure of the catalyst with a highly interconnected network of pores and channels.
- The scale bar in image d indicates 200 nm.
- These images likely represent different stages or variations of the iron-acetate/phenanthroline/zeolitic-imidazolate-framework-derived electrocatalyst mentioned in the article.
- The varying structures and porosities shown in these images could be related to the enhanced mass-transport properties described in the text.
- The varying structures and porosities in the images may contribute to increased volumetric activity described in the article.
"""

from pycocoevalcap.cider.cider import Cider
cider_scorer = Cider()


def compute_cider(predictions, references):
    """
    predictions: list of strings
    references: list of list of strings
    """
    return cider_scorer.compute_score([{i: [ref for ref in refs]} for i, refs in enumerate(references)],
                                      {i: [pred] for i, pred in enumerate(predictions)})[0]


def get_all_metrics_for_reference_based_metrics(
    references, candidates, image_id_list, reference_based_metrics_file,
):
    # Initialize the metrics
    existing_data = load(reference_based_metrics_file) if osp.exists(reference_based_metrics_file) else {}
    bleu_metric = evaluate.load("bleu")
    rouge_metric = evaluate.load("rouge")
    meteor_metric = evaluate.load("meteor")
    bertscore_metric = evaluate.load("bertscore")
    # cider_metric = evaluate.load("cider")

    # bleu_scores = [[], [], [], [], []]  # B1, B2, B3, B4, BLEU
    # rouge_scores = [[], [], [], []]  # ROUGE1, ROUGE2, ROUGEL, ROUGELSUM
    # meteor_scores = []
    # bertscore_scores = []

    # Calculate scores for each sample
    idx = 1
    print(f"Calculating metrics for {len(references)} samples")
    assert len(references) == len(candidates) == len(image_id_list)
    for ref, cand, image_id in tqdm(zip(references, candidates, image_id_list)):
        if not cand.strip():
            print(cand)
            continue

        default_bleu_score = {"bleu": 0.0, "precisions": [0.0, 0.0, 0.0, 0.0]}
        if image_id not in existing_data:
            existing_data[image_id] = {}
        bleu_score = existing_data.get(image_id, {}).get('bleu_score', default_bleu_score)

        if bleu_score == default_bleu_score:
            try:
                bleu_score = bleu_metric.compute(predictions=[cand], references=ref)
            except:
                bleu_score = default_bleu_score
            existing_data[image_id]['bleu_score'] = bleu_score

        default_rouge_score = {
            "rouge1": 0.0,
            "rouge2": 0.0,
            "rougeL": 0.0,
            "rougeLsum": 0.0,
        }
        rouge_score = existing_data.get(image_id, {}).get('rouge_score', default_rouge_score)
        if rouge_score == default_rouge_score:
            try:
                rouge_score = rouge_metric.compute(predictions=[cand], references=ref)
            except:
                rouge_score = default_rouge_score
            existing_data[image_id]['rouge_score'] = rouge_score

        # meteor score
        default_meteor_score = 0.0
        meteor_score = existing_data.get(image_id, {}).get('meteor_score', default_meteor_score)
        if meteor_score == default_meteor_score:
            try:
                meteor_score = meteor_metric.compute(predictions=[cand], references=ref)[
                    "meteor"
                ]
            except:
                meteor_score = default_meteor_score
            existing_data[image_id]['meteor_score'] = meteor_score

        # bertscore
        default_bertscore_score = {"f1": [0.0]}
        bertscore_score = existing_data.get(image_id, {}).get('bertscore_score', default_bertscore_score)
        if bertscore_score == default_bertscore_score:
            try:
                bertscore_score = bertscore_metric.compute(
                    predictions=[cand], references=ref, lang="en"
                )
            except:
                bertscore_score = default_bertscore_score
            existing_data[image_id]['bertscore_score'] = bertscore_score

        # cider score
        default_cider_score = 0.0
        cider_score = existing_data.get(image_id, {}).get('cider_score', default_cider_score)
        if cider_score == default_cider_score:
            try:
                # cider_score = cider_metric.compute(predictions=[cand], references=ref)["score"]
                cider_score = compute_cider([cand], [ref])
            except:
                cider_score = default_cider_score
            existing_data[image_id]['cider_score'] = cider_score

        if idx % 50 == 0:
            print(f"Saving 50 samples to {reference_based_metrics_file}")
            dump(existing_data, reference_based_metrics_file)

        idx += 1

    dump(existing_data, reference_based_metrics_file)
    print(f"Saved all samples to {reference_based_metrics_file}")

    return existing_data


def llm_openai_judge(prediction, reference, dimension, prompt, model):

    if isinstance(reference, list):
        reference = reference[0]
    assert isinstance(prediction, str)

    if dimension == "fluency":
        sys_prompt = prompt.format(Second=prediction)  # no reference
        usr_prompt = f"Caption:\n{prediction}\n"
    else:
        sys_prompt = prompt.format(Target=reference, Second=prediction)
        usr_prompt = f"Oracle Caption:\n{reference}\n\nSecond Caption:\n{prediction}\n\n"
    usr_prompt += f"What is the {dimension} score (1-5)? Return the score ONLY!"

    model.system_prompt = sys_prompt

    answer = 0
    all_responses = model.generate(usr_prompt)

    for response in all_responses.choices:
        response = response.message.content.strip()
        try:
            answer = int(response)
            break
        except:
            for s in ["1", "2", "3", "4", "5"]:
                if s in response:
                    answer = int(s)
                    break
    return answer


def g_eval_generate(evaluator, usr_prompt):
    return evaluator.generate(usr_prompt)


def get_all_metrics_for_g_eval_score(references, candidates, evaluator, tmp_file, nproc=4):
    # relevance scores
    assert len(references) == len(candidates)
    tups = []
    indices = range(len(references))
    for reference, prediction in tqdm(zip(references, candidates)):
        dimension = "relevance"
        if isinstance(reference, list):
            reference = reference[0]
        assert isinstance(prediction, str)

        sys_prompt = g_eval_prompt.format(Target=reference, Second=prediction)
        usr_prompt = f"Oracle Caption:\n{reference}\n\nSecond Caption:\n{prediction}\n\n"
        usr_prompt += f"What is the {dimension} score (1-5)? Return the score ONLY!"

        evaluator.system_prompt = sys_prompt
        tups.append((evaluator, usr_prompt))

    ans = {}
    if osp.exists(tmp_file):
        ans = load(tmp_file)
    ans = {k: v for k, v in ans.items() if evaluator.fail_msg not in str(v)}
    tups = [x for x, i in zip(tups, indices) if i not in ans]
    indices = [i for i in indices if i not in ans]

    if len(indices):
        _ = track_progress_rich(
            g_eval_generate,
            tups,
            nproc=nproc,
            chunksize=nproc,
            keys=indices,
            save=tmp_file,
        )
    ans = load(tmp_file)
    return ans


def merge_rating(refer_based_metrics_output_file_name, g_eval_metrics_output_file_name, fact_score_metrics_output_file):
    # Update metrics columns for each row
    refer_based_metrics_output_file = load(refer_based_metrics_output_file_name)
    g_eval_metrics_output_file = load(g_eval_metrics_output_file_name)
    # fact_score_metrics_output_file = load(fact_score_metrics_output_file)
    for idx, item in refer_based_metrics_output_file.iterrows():
        ref_based_metrics = eval(item['reference_based_metrics'])
        refer_based_metrics_output_file.at[idx, 'B2'] = ref_based_metrics['bleu_score']['precisions'][1] * 100
        refer_based_metrics_output_file.at[idx, 'RL'] = ref_based_metrics['rouge_score']['rougeL'] * 100
        refer_based_metrics_output_file.at[idx, 'M'] = ref_based_metrics['meteor_score'] * 100
        refer_based_metrics_output_file.at[idx, 'CD'] = ref_based_metrics['cider_score'] * 100
        refer_based_metrics_output_file.at[idx, 'BS'] = ref_based_metrics['bertscore_score']['f1'][0] * 100
        refer_based_metrics_output_file.at[idx, 'GE'] = g_eval_metrics_output_file.iloc[idx]['g_eval_metrics']
        # refer_based_metrics_output_file.at[idx, 'FS'] = eval(fact_score_metrics_output_file.iloc[idx]['fact_score_metrics'])['score']

    # df = refer_based_metrics_output_file
    # # Filter rows for the two settings
    # df_filtered = df[df['setting'].isin(['w/o.Abstract_w/o.Content', 'w.Abstract_w/o.Content'])]
    # # metrics = ['B2', 'RL', 'M', 'BS', 'CD', 'GE']
    # metrics = ['B2', 'RL', 'M', 'BS', 'GE']

    # # Group by subject and setting, averaging metrics
    # grouped = df_filtered.groupby(['subject', 'setting'])[metrics].mean()

    # # Pivot to get multi-level columns: top level is setting, sub-level is metric
    # result_df = grouped.unstack(level='setting')
    # # Swap levels so that top level is setting
    # result_df.columns = result_df.columns.swaplevel(0, 1)
    # # Ensure columns order matches required settings
    # result_df = result_df.reindex(columns=['w/o.Abstract_w/o.Content', 'w.Abstract_w/o.Content'], level=0)

    # # Reset index to have subject as a column
    # result_df = result_df.reset_index()
    # return result_df

    df = refer_based_metrics_output_file

    # metrics = ['B2', 'RL', 'M', 'BS', 'CD', 'FS', 'GE']
    metrics = ['B2', 'RL', 'M', 'BS', 'CD', 'GE']

    subject_df = df.groupby('subject')[metrics].mean().reset_index()

    category_df = df.groupby('category')[metrics].mean().reset_index()
    category_df.rename(columns={'category': 'subject'}, inplace=True)
    category_df['subject'] = 'CATEGORY_' + category_df['subject']

    overall_row = df[metrics].mean().to_frame().T
    overall_row.insert(0, 'subject', 'Overall')

    result_df = pd.concat([subject_df, category_df, overall_row], ignore_index=True)

    return result_df


def dump_multiindex_excel(df, path):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws.cell(row=1, column=1, value="subject")
    col_levels = df.columns.levels
    col_codes = df.columns.codes
    level0_labels = [col_levels[0][i] for i in col_codes[0]]
    level1_labels = [col_levels[1][i] for i in col_codes[1]]
    unique_level0 = list(dict.fromkeys(level0_labels))

    col_offset = 2
    start = col_offset
    for lvl0 in unique_level0:
        count = level0_labels.count(lvl0)
        end = start + count - 1
        ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
        ws.cell(row=1, column=start, value=lvl0)
        start = end + 1

    ws.cell(row=2, column=1, value="subject")
    for idx, label in enumerate(level1_labels):
        ws.cell(row=2, column=col_offset + idx, value=label)

    for r_idx, row in enumerate(df.itertuples(index=False), start=3):
        ws.cell(row=r_idx, column=1, value=row[0])  # subject
        for c_idx, val in enumerate(row[1:], start=2):
            ws.cell(row=r_idx, column=c_idx, value=val)

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    wb.save(path)


def extract_int(s, scales=[0, 5]):
    # Regular expression to find float numbers between 0 and 1, including 0 and 1
    for score in range(scales[0], scales[1] + 1):
        if str(score) in s:
            return int(score)
    return None


def get_score(
    model,
    content,
    gamma=10,
    scales=[0, 5],
    atomic_facts=None,
):

    total_decisions = []
    scores = []
    init_scores = []

    decisions = []
    for atom in atomic_facts:
        if not atom:
            continue
        atom = atom.strip()

        definition = f"How relevant is the generated caption to the given human-written caption for the figure? Determine whether the information in the generated caption is included or mentioned in the human-written caption. Respond with a score between {scales[0]} and {scales[1]}."

        prompt = "Human-written caption: {}\n\nGenerated caption: {}\n\nHow relevant is the generated caption to the given human-written caption? Respond with ONLY a score between {} and {}.".format(
            content, atom, scales[0], scales[1]
        )

        model.system_prompt = definition
        outputs = model.generate(prompt)

        generated_answer = outputs.lower().strip()

        generated_score = extract_int(generated_answer, scales=scales)

        if generated_score is None:
            generated_score = 0.0
        else:
            try:
                # normalize the score to 0-1
                generated_score = float(generated_score) / scales[1]
            except Exception as e:
                print("Error:", e)
                generated_score = 0.0

        print("Atom score:", generated_score)

        decisions.append({"atom": atom, "is_supported": generated_score})

    score = np.mean([d["is_supported"] for d in decisions])
    if gamma:
        init_scores.append(score)
        penalty = (
            1.0 if len(atomic_facts) > gamma else np.exp(1 - gamma / len(atomic_facts))
        )
        score = penalty * score

    total_decisions.append(decisions)
    scores.append(score)
    print("Final score:", score)

    out = {
        "score": np.mean(scores),
        "decisions": total_decisions,
        "num_facts_per_response": np.mean(
            [len(d) for d in total_decisions if d is not None]
        ),
    }

    if gamma:
        out["init_score"] = np.mean(init_scores)

    return out


def fact_score_generate(model, line):
    generated_caption = line['prediction']
    model.system_prompt = generate_atomic_facts_sys_prompt
    mes = [
        dict(type='text', value=generate_atomic_facts_usr_prompt_one_shot),
        dict(type='text', value=generated_caption + '\nModel_Response:'),
    ]
    generated_facts = model.generate(mes).strip()
    formatted_facts = generated_facts.split("-")
    # topic = line["subject"]
    # abstract = line["abstract"]

    # ground-truth caption
    content = line["caption"]

    score_out = get_score(
        model,
        content,
        gamma=10,
        scales=[0, 5],
        atomic_facts=formatted_facts,
    )
    return score_out
